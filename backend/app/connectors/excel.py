"""Excel (.xlsx) connector.

Reads sheets from a local .xlsx / .xlsm file via openpyxl (pure Python,
air-gap safe — no runtime extension downloads). Each sheet becomes a
queryable table whose columns are taken from the first row. SQL runs
against an in-memory DuckDB that's loaded fresh on every query, so
there's no caching layer to invalidate.

Use cases:
  - A customer hands you an Excel file from their analyst.
  - One-off analysis on a spreadsheet exported from another tool.
  - Demos / training without setting up a database.

Limitations (acceptable for v1):
  - Whole file is read into memory on every query. Fine up to ~100K rows
    per sheet; revisit if customers ask about huge spreadsheets.
  - No type inference beyond what openpyxl gives us (numbers, dates,
    strings stay as-is; everything else stringifies).
  - First row is assumed to be the header. Blank header cells become
    'column_<n>'.
"""

from __future__ import annotations

import asyncio
import re
from pathlib import Path

import duckdb
import openpyxl

from app.connectors.base import (
    Connector,
    ConnectorError,
    IntrospectedColumn,
    IntrospectedTable,
    QueryResult,
)

# Reject anything that looks like a write — DuckDB CAN write to in-memory
# tables, but for a read-only product surface we keep mutations out.
_WRITE_STATEMENTS = re.compile(
    r"^\s*(insert|update|delete|drop|alter|create|truncate|attach|copy|export)\b",
    re.IGNORECASE,
)


def _safe_identifier(name: str) -> str:
    """Sanitise a sheet name into a usable SQL identifier.

    DuckDB is permissive with quoted identifiers but the LLM is happier
    with bare names. Strip everything that isn't alphanumeric / underscore
    and prefix with `t_` if it would start with a digit.
    """
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", name).strip("_")
    if not cleaned:
        cleaned = "sheet"
    if cleaned[0].isdigit():
        cleaned = "t_" + cleaned
    return cleaned


def _coerce_header(value, position: int) -> str:  # noqa: ANN001
    raw = "" if value is None else str(value).strip()
    return raw or f"column_{position}"


class ExcelConnector(Connector):
    type = "excel"
    display_name = "Excel (.xlsx)"
    credential_keys = frozenset()  # local file, no secrets

    def _path(self) -> Path:
        raw = (self.config or {}).get("path")
        if not raw:
            raise ConnectorError("Excel source requires a 'path' config field")
        return Path(str(raw)).expanduser()

    # ── Lifecycle ────────────────────────────────────────────────────

    async def healthcheck(self) -> bool:
        def _check() -> bool:
            try:
                p = self._path()
            except ConnectorError:
                return False
            if not p.exists() or not p.is_file():
                return False
            if p.suffix.lower() not in {".xlsx", ".xlsm"}:
                return False
            # Cheap parse — just open and close the workbook.
            try:
                wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                wb.close()
                return True
            except Exception:
                return False

        return await asyncio.to_thread(_check)

    async def introspect(self) -> list[IntrospectedTable]:
        def _introspect() -> list[IntrospectedTable]:
            p = self._path()
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            tables: list[IntrospectedTable] = []
            try:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows_iter = ws.iter_rows(values_only=True)
                    try:
                        header_row = next(rows_iter)
                    except StopIteration:
                        # Empty sheet — still expose it but with no columns.
                        tables.append(
                            IntrospectedTable(
                                schema_name="main",
                                name=_safe_identifier(sheet_name),
                                description=f"Sheet: {sheet_name} (empty)",
                            )
                        )
                        continue

                    columns: list[IntrospectedColumn] = []
                    for i, cell in enumerate(header_row):
                        columns.append(
                            IntrospectedColumn(
                                name=_coerce_header(cell, i),
                                data_type="text",
                                nullable=True,
                                is_primary_key=False,
                                position=i,
                                description=None,
                            )
                        )

                    # Row count estimate — cheap when read_only=True (uses
                    # the sheet dimensions, not a full scan).
                    row_est: int | None = ws.max_row - 1 if ws.max_row else None
                    if row_est is not None and row_est < 0:
                        row_est = 0

                    tables.append(
                        IntrospectedTable(
                            schema_name="main",
                            name=_safe_identifier(sheet_name),
                            description=f"Sheet: {sheet_name}",
                            row_count_est=row_est,
                            columns=columns,
                        )
                    )
            finally:
                wb.close()
            return tables

        return await asyncio.to_thread(_introspect)

    async def estimate_rows(self, sql: str) -> int | None:
        # We could spin up the in-memory DuckDB and EXPLAIN against it,
        # but it's the same cost as just running the query. Skip for now;
        # the orchestrator handles None gracefully.
        return None

    async def execute(
        self,
        sql: str,
        *,
        max_rows: int = 10_000,
        timeout_seconds: int = 60,
    ) -> QueryResult:
        if _WRITE_STATEMENTS.search(sql):
            raise ConnectorError(
                "write statements are not allowed via the Excel connector "
                "(read-only product surface)."
            )

        def _run() -> QueryResult:
            p = self._path()
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            con = duckdb.connect(database=":memory:")
            try:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows_iter = ws.iter_rows(values_only=True)
                    try:
                        header_row = next(rows_iter)
                    except StopIteration:
                        continue
                    headers = [_coerce_header(c, i) for i, c in enumerate(header_row)]
                    data = list(rows_iter)
                    table_name = _safe_identifier(sheet_name)
                    # CREATE TABLE then INSERT — stays in stdlib types, no
                    # pandas required. Everything stored as VARCHAR; the LLM
                    # can CAST in queries when numeric ops are needed.
                    quoted_cols = ", ".join(f'"{h}" VARCHAR' for h in headers)
                    con.execute(f'CREATE TABLE "{table_name}" ({quoted_cols})')
                    if data:
                        placeholders = ", ".join("?" for _ in headers)
                        # Stringify everything for storage simplicity; DuckDB
                        # can CAST in queries when the LLM writes numeric ops.
                        rows_str = [
                            tuple("" if v is None else str(v) for v in row)
                            for row in data
                        ]
                        con.executemany(
                            f'INSERT INTO "{table_name}" VALUES ({placeholders})',
                            rows_str,
                        )

                cur = con.execute(sql)
                fetched = cur.fetchall()
                columns = [d[0] for d in cur.description] if cur.description else []
            finally:
                con.close()
                wb.close()

            truncated = len(fetched) > max_rows
            limited = fetched[:max_rows]
            return QueryResult(
                columns=columns,
                rows=limited,
                row_count=len(limited),
                truncated=truncated,
            )

        try:
            return await asyncio.to_thread(_run)
        except duckdb.Error as exc:
            raise ConnectorError(f"query failed: {exc}") from exc
